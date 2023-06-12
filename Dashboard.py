from tkinter import *
import openpyxl
# import xlrd
from openpyxl import Workbook
from tkinter.ttk import Combobox
from tkinter import messagebox, ttk
import pathlib
from PIL import Image, ImageTk


class KTP(Toplevel):

    def __init__(self, master=None):

        super().__init__(master=master)
        self.title("New Window")
        self.geometry("500x320")
        self.canfasr = Frame(self)
        self.canfasr.place(x=0, y=0)
        self.canfasr.configure(height=500, width=320)
        # label = Label(self, text="This is a new Window")
        self.bg = ImageTk.PhotoImage(file="ktp.png")
        self.canvas = Canvas(self.canfasr, width=500, height=320)
        self.canvas.pack(fill="both", expand=False)
        # Display image
        self.canvas.create_image(0, 0, image=self.bg, anchor="nw")
        # label.pack()

    def cetak_ktp(baris):
        print(baris)


def click(e):
    item = tree.selection()[0]
    # print("you clicked on", int(tree.item(item, "text"))+1)
    # print(int(tree.item(item, "text")))
    KTP(root)
    KTP.cetak_ktp(int(tree.item(item, "text"))+1)


def data_cek(dat):
    list_data = dat
    for i in list_data:
        if i == '':
            return False
            break
    else:
        return True


def submit():

    list_data = [nik.get(), nama.get(), tempat_lahir.get(), tanggal_lahir.get(), bulan_lahir.get(), tahun_lahir.get(), alamat.get(), rt.get(), rw.get(), kelurahan.get(
    ), kecamatan.get(), kabupaten.get(), provinsi.get(), jenis_kelamin.get(), goldarah.get(), pekerjaan.get(), agama.get(), status.get(), kewarganegaraan.get()]
    cek_data = data_cek(list_data)

    if cek_data == False:
        messagebox.showinfo('error', 'Data belum lengkap')
    else:
        file = openpyxl.load_workbook('Data_Base.xlsx')
        sheet = file.active
        baris = sheet.max_row+1
        kolom = 0
        # entri data ke data base
        file = openpyxl.load_workbook('Data_Base.xlsx')
        sheet = file.active
        for i in list_data:
            kolom += 1
            sheet.cell(column=kolom, row=baris, value=i)
        file.save('Data_Base.xlsx')
        messagebox.showinfo('info', 'data tersimpan!!!')
        clear_entry()


def clear_entry():
    nik.set('')
    nama.set('')
    tempat_lahir.set('')
    tanggal_lahir.set('')
    bulan_lahir.set('')
    tahun_lahir.set('')
    alamat.set('')
    rt.set('')
    rw.set('')
    kelurahan.set('')
    kecamatan.set('')
    kabupaten.set('')
    provinsi.set('')
    jenis_kelamin.set('')
    goldarah.set('')
    pekerjaan.set('')
    agama.set('')
    status.set('')
    kewarganegaraan.set('')


def page_tambah():
    # label
    Label(window, text='Nama', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=0, padx=5, pady=5)
    Label(window, text='NIK', font=20, fg='#000000',
          bg='#eff5f6').grid(column=4, row=0, padx=5, pady=5)
    Label(window, text='Tempat Lahir', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=1, padx=5, pady=5)
    Label(window, text='Tanggal', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=2, padx=5, pady=5)
    Label(window, text='Bulan', font=20, fg='#000000',
          bg='#eff5f6').grid(column=2, row=2, padx=5, pady=5)
    Label(window, text='Tahun', font=20, fg='#000000',
          bg='#eff5f6').grid(column=4, row=2, padx=5, pady=5)
    Label(window, text='Jenis Kelamin', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=3, padx=5, pady=5)
    Label(window, text='Golongan Darah', font=20, fg='#000000',
          bg='#eff5f6').grid(column=2, row=3, padx=5, pady=5)
    Label(window, text='Alamat', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=4, padx=5, pady=5)
    Label(window, text='RT', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=5, padx=5, pady=5)
    Label(window, text='RW', font=20, fg='#000000',
          bg='#eff5f6').grid(column=2, row=5, padx=5, pady=5)
    Label(window, text='Kel/Desa', font=20, fg='#000000',
          bg='#eff5f6').grid(column=4, row=5, padx=5, pady=5)
    Label(window, text='Kecamatan', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=6, padx=5, pady=5)
    Label(window, text='Kabupaten/Kota', font=20, fg='#000000',
          bg='#eff5f6').grid(column=2, row=6, padx=5, pady=5)
    Label(window, text='Provinsi', font=20, fg='#000000',
          bg='#eff5f6').grid(column=4, row=6, padx=5, pady=5)
    Label(window, text='Agama', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=7, padx=5, pady=5)
    Label(window, text='Kewarga Negaraan', font=20, fg='#000000',
          bg='#eff5f6').grid(column=2, row=7, padx=5, pady=5)
    Label(window, text='Status Perkawinan', font=20, fg='#000000',
          bg='#eff5f6').grid(column=4, row=7, padx=5, pady=5)
    Label(window, text='Pekerjaan', font=20, fg='#000000',
          bg='#eff5f6').grid(column=0, row=8, padx=5, pady=5)

    # tambah_frame = Frame(window)

    # Entry
    entri_nama = Entry(window, textvariable=nama, width=80,).grid(
        column=1, row=0, columnspan=3, sticky=W, padx=5, pady=5)
    entri_nik = Entry(window, textvariable=nik, width=20).grid(
        column=5, row=0, sticky=W, padx=5, pady=5)
    entri_tl = Entry(window, textvariable=tempat_lahir, width=30).grid(
        column=1, row=1, sticky=W, padx=5, pady=5)
    entri_tgl = Entry(window, textvariable=tanggal_lahir, width=5).grid(
        column=1, row=2, sticky=W, padx=5, pady=5)
    entri_bln = Entry(window, textvariable=bulan_lahir, width=5).grid(
        column=3, row=2, sticky=W, padx=5, pady=5)
    entri_thn = Entry(window, textvariable=tahun_lahir, width=5).grid(
        column=5, row=2, sticky=W, padx=5, pady=5)
    entry_jk = Combobox(window,  values=['LAKI-LAKI', 'PEREMPUAN'], textvariable=jenis_kelamin,
                        state='r').grid(column=1, row=3, sticky=W, padx=5, pady=5)
    entri_golD = Combobox(window,  values=['A', 'B', 'O', 'AB'], textvariable=goldarah, state='r').grid(
        column=3, row=3, sticky=W, padx=5, pady=5)
    entri_almt = Entry(window, textvariable=alamat, width=20).grid(
        column=1, row=4, sticky=W, padx=5, pady=5)
    entri_rt = Entry(window, textvariable=rt, width=5).grid(
        column=1, row=5, sticky=W, padx=5, pady=5)
    entri_tw = Entry(window, textvariable=rw, width=5).grid(
        column=3, row=5, sticky=W, padx=5, pady=5)
    entri_kel = Entry(window, textvariable=kelurahan, width=20).grid(
        column=5, row=5, sticky=W, padx=5, pady=5)
    entri_kec = Entry(window, textvariable=kecamatan, width=20).grid(
        column=1, row=6, sticky=W, padx=5, pady=5)
    entri_kab = Entry(window, textvariable=kabupaten, width=20).grid(
        column=3, row=6, sticky=W, padx=5, pady=5)
    entri_prov = Entry(window, textvariable=provinsi, width=20).grid(
        column=5, row=6, sticky=W, padx=5, pady=5)
    entri_agm = Combobox(window, textvariable=agama, values=[
                         'ISLAM', 'KRISTEN', 'KATOLIK', 'HINDU', 'BUDHA', 'KONGHUCHU'], state='r').grid(column=1, row=7, sticky=W, padx=5, pady=5)
    entri_status = Combobox(window, textvariable=status, values=[
                            'KAWIN', 'BELUM KAWIN'], state='r').grid(column=3, row=7, sticky=W, padx=5, pady=5)
    entri_kwn = Combobox(window, textvariable=kewarganegaraan, values=[
                         'WNI', 'WNA'], state='r').grid(column=5, row=7, sticky=W, padx=5, pady=5)
    entri_kerja = Entry(window, textvariable=pekerjaan, width=20).grid(
        column=1, row=8, sticky=W, padx=5, pady=5)

    # submitbutton
    Button(window, text='clear', bg='#009df4', fg='black', width=10,
           font=("", 10, "bold"), command=clear_entry).grid(column=0, row=9)
    Button(window, text='Submit', bg='#009df4', fg='white', width=10, height=2, font=(
        "", 13, "bold"), command=submit).grid(column=0, row=10, columnspan=6)


def page_lihat():
    scrol = Scrollbar(window, orient='vertical')
    global tree
    tree = ttk.Treeview(window, yscrollcommand=scrol.set,
                        column=('c0', 'c1', 'c3'), height=20)

    tree.grid(row=3, column=0, columnspan=2)
    tree.column('#0', width=30, anchor=W, stretch=NO)
    tree.column('#1', width=300, anchor=W, stretch=NO)
    tree.column('#2', width=670, anchor=W, stretch=NO)

    tree.heading('#0', text='NO')
    tree.heading('#1', text='NIK')
    tree.heading('#2', text='Nama')

    file = openpyxl.load_workbook('Data_Base.xlsx')
    sheet = file.active
    # menampilkan data
    for i in range(2, sheet.max_row+1):
        nik = sheet.cell(column=1, row=i)
        nama = sheet.cell(column=2, row=i)
        tree.insert('', 'end', text=f'{i-1}',
                    values=(f'{nik.value}', f'{nama.value}'))
    # saat baris di click
    tree.bind("<Double-1>", click)
    lihat_data = Label(root, text='Double click untuk menampilkan KTP', font=20, fg='#000000',
                       bg='#eff5f6')
    lihat_data.place(y=500, relx=0.4)
    scrol.configure(command=tree.yview)
    scrol.grid(row=0, column=3, rowspan=100, sticky=NS)
    file.save('Data_Base.xlsx')


def page_edit():
    scrol = Scrollbar(window, orient='vertical')
    global tree
    tree = ttk.Treeview(window, yscrollcommand=scrol.set,
                        column=('c0', 'c1', 'c3'))
    tree.grid(row=0, column=0, rowspan=100, columnspan=2)
    tree.column('#0', width=30, anchor=W, stretch=NO)
    tree.column('#1', width=300, anchor=W, stretch=NO)
    tree.column('#2', width=670, anchor=W, stretch=NO)

    tree.heading('#0', text='NO')
    tree.heading('#1', text='NIK')
    tree.heading('#2', text='Nama')

    file = openpyxl.load_workbook('Data_Base.xlsx')
    sheet = file.active
    # menampilkan data
    for i in range(2, sheet.max_row+1):
        nik = sheet.cell(column=1, row=i)
        nama = sheet.cell(column=2, row=i)
        tree.insert('', 'end', text=f'{i-1}',
                    values=(f'{nik.value}', f'{nama.value}'))

    # saat baris di click
    tree.bind("<Double-1>", click)

    scrol.configure(command=tree.yview)
    scrol.grid(row=0, column=3, rowspan=100, sticky=NS)
    file.save('Data_Base.xlsx')


def page_hapus():
    scrol = Scrollbar(window, orient='vertical')
    global tree
    tree = ttk.Treeview(window, yscrollcommand=scrol.set,
                        column=('c0', 'c1', 'c3'))
    tree.grid(row=0, column=0, rowspan=100, columnspan=2)
    tree.column('#0', width=30, anchor=W, stretch=NO)
    tree.column('#1', width=300, anchor=W, stretch=NO)
    tree.column('#2', width=670, anchor=W, stretch=NO)

    tree.heading('#0', text='NO')
    tree.heading('#1', text='NIK')
    tree.heading('#2', text='Nama')

    file = openpyxl.load_workbook('Data_Base.xlsx')
    sheet = file.active
    # menampilkan data
    for i in range(2, sheet.max_row+1):
        nik = sheet.cell(column=1, row=i)
        nama = sheet.cell(column=2, row=i)
        tree.insert('', 'end', text=f'{i-1}',
                    values=(f'{nik.value}', f'{nama.value}'))

    # saat baris di click
    tree.bind("<Double-1>", click)

    scrol.configure(command=tree.yview)
    scrol.grid(row=0, column=3, rowspan=100, sticky=NS)
    file.save('Data_Base.xlsx')


def delete_frame():
    for frame in window.winfo_children():
        frame.destroy()


def hapus_indikator():
    tambah_data.config(bg='#ffffff', fg='#000000')
    lihat_data.config(bg='#ffffff', fg='#000000')
    edit_data.config(bg='#ffffff', fg='#000000')
    hapus_data.config(bg='#ffffff', fg='#000000')


def indikator(lb, page):
    hapus_indikator()
    lb.config(bg='#009df4', fg='white')
    delete_frame()
    page()


if __name__ == '__main__':
    root = Tk()
    # inisialisasi variabel
    nik = StringVar()
    nama = StringVar()
    tempat_lahir = StringVar()
    tanggal_lahir = StringVar()
    bulan_lahir = StringVar()
    tahun_lahir = StringVar()
    alamat = StringVar()
    rt = StringVar()
    rw = StringVar()
    kelurahan = StringVar()
    kecamatan = StringVar()
    kabupaten = StringVar()
    provinsi = StringVar()
    jenis_kelamin = StringVar()
    goldarah = StringVar()
    pekerjaan = StringVar()
    agama = StringVar()
    status = StringVar()
    kewarganegaraan = StringVar()

    entri_nama = ''

    # cek Database
    file = pathlib.Path('Data_Base.xlsx')
    if file.exists():
        pass
    else:
        file = Workbook()
        sheet = file.active
        sheet['A1'] = 'NIK'
        sheet['B1'] = 'Nama'
        sheet['C1'] = 'tempat_lahir'
        sheet['D1'] = 'tanggal_lahir'
        sheet['E1'] = 'bullan_lahir'
        sheet['F1'] = 'tahun_lahir'
        sheet['G1'] = 'alamat'
        sheet['H1'] = 'rt'
        sheet['I1'] = 'rw'
        sheet['J1'] = 'kelurahan'
        sheet['K1'] = 'kecamatan'
        sheet['L1'] = 'kabupater'
        sheet['M1'] = 'provinsi'
        sheet['N1'] = 'jenis_kelamin'
        sheet['O1'] = 'gol_darah'
        sheet['P1'] = 'pekerjaan'
        sheet['Q1'] = 'agama'
        sheet['R1'] = 'status'
        sheet['S1'] = 'kewarganegaraan'

        file.save('Data_Base.xlsx')

    root.title('Program E-KTP')
    root.geometry('1200x600')
    root.resizable(0, 0)
    root.config(background='#eff5f6')

    # Header
    header = Frame(root, bg='#009df4')
    header.place(x=200, y=0, width=1000, height=60)
    label = Label(header, text="E-KTP", bg='#009df4',
                  font=("", 13, "bold"), bd=0, fg='white')
    label.place(relx=0.5, rely=0.5, anchor='center')

    # side_bar
    side_bar = Frame(root, bg='#ffffff')
    side_bar.place(x=0, y=0, width=200, height=600)

    # Menu
    tambah_data = Button(side_bar, text='Tambah Data', font=("", 12, "bold"), bg='#ffffff',
                         bd=0, cursor='hand2', command=lambda: indikator(tambah_data, page_tambah))
    tambah_data.place(y=100, width=200, height=60)
    # tambah_data.indikator = Label(tambah_data, bg='white')
    lihat_data = Button(side_bar, text='Lihat Data', font=("", 12, "bold"), bg='#ffffff',
                        bd=0, cursor='hand2', command=lambda: indikator(lihat_data, page_lihat))
    lihat_data.place(y=160, width=200, height=60)
    edit_data = Button(side_bar, text='Edit Data', font=("", 12, "bold"), bg='#ffffff',
                       bd=0, cursor='hand2', command=lambda: indikator(edit_data, page_edit))
    edit_data.place(y=220, width=200, height=60)
    hapus_data = Button(side_bar, text='Hapus Data', font=("", 12, "bold"), bg='#ffffff',
                        bd=0, cursor='hand2', command=lambda: indikator(hapus_data, page_hapus))
    hapus_data.place(y=280, width=200, height=60)

    # main window
    window = Frame(root)
    window.place(x=200, y=60)
    window.config(bg='#eff5f6')
    # window.pack_propagate(False)
    window.configure(height=540, width=1000)

    root.mainloop()
